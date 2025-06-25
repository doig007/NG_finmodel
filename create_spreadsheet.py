import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuration & Constants ---
# Years for historical data to be manually entered or linked if available in digital form
HISTORICAL_YEARS_DATA = ["FY2020", "FY2021", "FY2022", "FY2023", "FY2024"]
# Years for forecast period
FORECAST_YEARS_MODEL = [f"FY{2025 + i}" for i in range(16)] # FY2025 to FY2040
# Columns to display on sheets that show both history and forecast
DISPLAY_YEARS = HISTORICAL_YEARS_DATA[-2:] + FORECAST_YEARS_MODEL # Last 2 historical + all forecast

# --- Styling Definitions ---
# Colors (Hex format)
COLOR_PRIMARY_BLUE = "4F81BD" # Darker Blue (Ofgem/NatGrid style)
COLOR_SECONDARY_BLUE = "DCE6F1" # Lighter Blue
COLOR_WHITE = "FFFFFF"
COLOR_BLACK = "000000"
COLOR_INPUT_BLUE = "0000FF" # Standard input blue
COLOR_GREEN_LINK = "008000" # Green for links (conceptual)
COLOR_GREY_FILL = "F2F2F2" # Light grey for some subheaders or read-only calc sections

# Fonts
FONT_HEADER = Font(bold=True, color=COLOR_WHITE, name='Calibri', size=11)
FONT_SUBHEADER = Font(bold=True, color=COLOR_BLACK, name='Calibri', size=11)
FONT_INPUT = Font(color=COLOR_INPUT_BLUE, name='Calibri', size=10)
FONT_FORMULA = Font(color=COLOR_BLACK, name='Calibri', size=10)
FONT_LINK = Font(color=COLOR_GREEN_LINK, name='Calibri', size=10) # For conceptual marking

# Fills
FILL_HEADER = PatternFill(start_color=COLOR_PRIMARY_BLUE, end_color=COLOR_PRIMARY_BLUE, fill_type="solid")
FILL_SUBHEADER = PatternFill(start_color=COLOR_SECONDARY_BLUE, end_color=COLOR_SECONDARY_BLUE, fill_type="solid")
FILL_GREY = PatternFill(start_color=COLOR_GREY_FILL, end_color=COLOR_GREY_FILL, fill_type="solid")

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Borders
BORDER_THIN_ALL = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
BORDER_BOTTOM_MEDIUM = Border(bottom=Side(style='medium'))

# Number Formats
FORMAT_PERCENT_1DP = '0.0%'
FORMAT_PERCENT_0DP = '0%'
FORMAT_NUMBER_0DP = '#,##0'
FORMAT_NUMBER_0DP_NEG_PAREN = '#,##0;(#,##0)'
FORMAT_NUMBER_1DP = '#,##0.0'
FORMAT_NUMBER_2DP = '#,##0.00'
FORMAT_MULTIPLIER = '0.0x'

# --- Helper Functions ---
def setup_sheet_headers(ws, title, years_list, first_data_col_idx=2, row_num=1, main_header_fill=FILL_HEADER, year_header_fill=FILL_HEADER, notes_col=True):
    """Sets up the main title and year headers for a sheet."""
    ws.cell(row=row_num, column=1, value=title).font = FONT_SUBHEADER # Title in first column usually
    ws.cell(row=row_num, column=1).border = BORDER_THIN_ALL
    
    for i, year in enumerate(years_list):
        cell = ws.cell(row=row_num, column=first_data_col_idx + i, value=year)
        cell.font = FONT_HEADER
        cell.fill = year_header_fill
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN_ALL
    if notes_col:
        notes_col_idx = first_data_col_idx + len(years_list)
        cell = ws.cell(row=row_num, column=notes_col_idx, value="Notes/Links")
        cell.font = FONT_HEADER
        cell.fill = year_header_fill
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN_ALL

def style_row_header(cell, level=1):
    """Styles a row header cell."""
    cell.font = FONT_SUBHEADER if level == 1 else Font(bold=True, name='Calibri', size=10)
    if level == 1:
        cell.fill = FILL_SUBHEADER
    cell.alignment = ALIGN_LEFT
    cell.border = BORDER_THIN_ALL

def style_data_cell(cell, is_input=False, is_link=False, number_format=FORMAT_NUMBER_0DP_NEG_PAREN):
    """Styles a data cell, distinguishing inputs."""
    if is_input:
        cell.font = FONT_INPUT
    # elif is_link: # Conceptual, as actual link styling is more complex
    #     cell.font = FONT_LINK
    else:
        cell.font = FONT_FORMULA
    cell.number_format = number_format
    cell.alignment = ALIGN_RIGHT
    cell.border = BORDER_THIN_ALL

def set_column_widths(ws, widths_dict):
    """Sets column widths from a dictionary {col_letter: width}."""
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width

# --- Create Workbook ---
wb = openpyxl.Workbook()
wb.remove(wb.active) # Remove default sheet

# --- Sheet: Assumptions_Macro ---
ws_am = wb.create_sheet("Assumptions_Macro")
set_column_widths(ws_am, {'A': 45, **{get_column_letter(i+2): 12 for i in range(len(FORECAST_YEARS_MODEL))}, get_column_letter(len(FORECAST_YEARS_MODEL)+2): 50})
setup_sheet_headers(ws_am, "Macro & Group Assumptions", FORECAST_YEARS_MODEL)

macro_data = [
    ("MACROECONOMIC", None, None, "", True), # Item, Values, NumFormat, Notes, IsSectionHeader
    ("UK CPIH (Annual %)", [0.03, 0.025] + [0.02]*14, FORMAT_PERCENT_1DP, "Illustrative path to long-term target"),
    ("US CPI (Annual %)", [0.028, 0.023] + [0.02]*14, FORMAT_PERCENT_1DP, "Illustrative path to long-term target"),
    ("UK 10-yr Gilt Yield (Avg %)", [0.042, 0.038, 0.035] + [0.035]*13, FORMAT_PERCENT_1DP, "For cost of debt estimates"),
    ("US 10-yr Treasury Yield (Avg %)", [0.043, 0.040, 0.038] + [0.038]*13, FORMAT_PERCENT_1DP, "For cost of debt estimates"),
    ("GBP:USD Exchange Rate (Average)", [1.25, 1.26, 1.27] + [1.28]*13, FORMAT_NUMBER_2DP, "For P&L/CF translation"),
    ("GBP:USD Exchange Rate (Year End)", [1.26, 1.27, 1.28] + [1.29]*13, FORMAT_NUMBER_2DP, "For BS translation"),
    ("FINANCING & GROUP", None, None, "", True),
    ("Cost of New Debt (GBP %)", [0.055, 0.053] + [0.050]*14, FORMAT_PERCENT_1DP, "Illustrative, spread over Gilts"),
    ("Cost of New Debt (USD %)", [0.058, 0.055] + [0.053]*14, FORMAT_PERCENT_1DP, "Illustrative, spread over Treasuries"),
    ("UK Corporation Tax Rate (%)", [0.25]*16, FORMAT_PERCENT_0DP, "Current legislated"),
    ("US Federal Corp Tax Rate (%)", [0.21]*16, FORMAT_PERCENT_0DP, "Current legislated"),
    ("US Blended State Tax (Net of Fed Benefit, %)", [0.04]*16, FORMAT_PERCENT_1DP, "Illustrative effective state rate"),
    ("Dividend Payout Ratio (% of Net Profit to Equity Holders)", [0.60]*16, FORMAT_PERCENT_0DP, "Illustrative policy"),
    ("Target Minimum Cash Balance (£m)", [1000]*16, FORMAT_NUMBER_0DP, "Operational liquidity target"),
    ("Number of Shares Outstanding (millions)", [3700]*16, FORMAT_NUMBER_0DP, "Illustrative, for EPS calc; assumes no buybacks/issuance")
]
current_row = 2
for item_data in macro_data:
    item, values, num_format, notes_text = item_data[0], item_data[1], item_data[2], item_data[3]
    is_section_header = item_data[4] if len(item_data) > 4 else False
    cell_A = ws_am.cell(row=current_row, column=1, value=item)
    if is_section_header:
        ws_am.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(FORECAST_YEARS_MODEL)+2)
        cell_A.font = FONT_HEADER
        cell_A.fill = FILL_HEADER
        cell_A.alignment = ALIGN_CENTER
    else:
        style_row_header(cell_A, level=2)
        ws_am.cell(row=current_row, column=len(FORECAST_YEARS_MODEL)+2, value=notes_text).border = BORDER_THIN_ALL
        for i, val in enumerate(values):
            data_cell = ws_am.cell(row=current_row, column=i+2, value=val)
            style_data_cell(data_cell, is_input=True, number_format=num_format)
    current_row += 1

# --- Sheet: Hist_PL_Segment ---
ws_hpl = wb.create_sheet("Hist_PL_Segment")
set_column_widths(ws_hpl, {'A': 40, **{get_column_letter(i+2): 12 for i in range(len(HISTORICAL_YEARS_DATA))}, get_column_letter(len(HISTORICAL_YEARS_DATA)+2): 40})
setup_sheet_headers(ws_hpl, "Historical P&L by Segment (£m)", HISTORICAL_YEARS_DATA)
hpl_rows = [ # (Description, FY20, FY21, FY22, FY23, FY24, Notes, Is_Header, Num_Format)
    ("UK Electricity Transmission (NGET)", None, None, None, None, None, "", True, None),
    ("Revenue", 3000, 3100, 3200, 3300, 3400, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Operating Costs", -1000, -1050, -1100, -1150, -1200, "Excludes D&A", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("EBITDA", "=B3+B4", "=C3+C4", "=D3+D4", "=E3+E4", "=F3+F4", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Depreciation & Amort.", -500, -520, -540, -560, -580, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Operating Profit (EBIT)", "=B5+B6", "=C5+C6", "=D5+D6", "=E5+E6", "=F5+F6", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("UK Electricity Distribution (NGED)", None, None, None, None, None, "", True, None),
    ("Revenue", 0, 0, 1000, 3500, 3600, "Acquired part way FY22", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Operating Costs", 0, 0, -300, -1200, -1250, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("EBITDA", "=B9+B10", "=C9+C10", "=D9+D10", "=E9+E10", "=F9+F10", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Depreciation & Amort.", 0, 0, -150, -600, -620, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Operating Profit (EBIT)", "=B11+B12", "=C11+C12", "=D11+D12", "=E11+E12", "=F11+F12", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("US Regulated", None, None, None, None, None, "", True, None),
    ("Revenue", 5000, 5200, 5400, 5600, 5800, "FX impact incl.", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Operating Costs", -2500, -2600, -2700, -2800, -2900, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("EBITDA", "=B15+B16", "=C15+C16", "=D15+D16", "=E15+E16", "=F15+F16", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Depreciation & Amort.", -1000, -1050, -1100, -1150, -1200, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Operating Profit (EBIT)", "=B17+B18", "=C17+C18", "=D17+D18", "=E17+E18", "=F17+F18", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("National Grid Ventures (NGV)", None, None, None, None, None, "", True, None),
    ("Revenue", 800, 850, 900, 950, 1000, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Operating Costs", -300, -320, -340, -360, -380, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("EBITDA", "=B21+B22", "=C21+C22", "=D21+D22", "=E21+E22", "=F21+F22", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Depreciation & Amort.", -100, -110, -120, -130, -140, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Operating Profit (EBIT)", "=B23+B24", "=C23+C24", "=D23+D24", "=E23+E24", "=F23+F24", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Other / Eliminations", None, None, None, None, None, "", True, None),
    ("Revenue", -100, -100, -100, -100, -100, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Operating Costs", -200, -210, -220, -230, -240, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("EBITDA", "=B27+B28", "=C27+C28", "=D27+D28", "=E27+E28", "=F27+F28", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Depreciation & Amort.", -50, -50, -50, -50, -50, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Operating Profit (EBIT)", "=B29+B30", "=C29+C30", "=D29+D30", "=E29+E30", "=F29+F30", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("GROUP CONSOLIDATED", None, None, None, None, None, "", True, None),
    ("Total Revenue", "=SUM(B3,B9,B15,B21,B27)", "=SUM(C3,C9,C15,C21,C27)", "=SUM(D3,D9,D15,D21,D27)", "=SUM(E3,E9,E15,E21,E27)", "=SUM(F3,F9,F15,F21,F27)", "Sum of segments", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Total EBITDA", "=SUM(B5,B11,B17,B23,B29)", "=SUM(C5,C11,C17,C23,C29)", "=SUM(D5,D11,D17,D23,D29)", "=SUM(E5,E11,E17,E23,E29)", "=SUM(F5,F11,F17,F23,F29)", "Sum of segments", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Total Deprec. & Amort.", "=SUM(B6,B12,B18,B24,B30)", "=SUM(C6,C12,C18,C24,C30)", "=SUM(D6,D12,D18,D24,D30)", "=SUM(E6,E12,E18,E24,E30)", "=SUM(F6,F12,F18,F24,F30)", "Sum of segments", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Total Operating Profit (EBIT)", "=B34+B35", "=C34+C35", "=D34+D35", "=E34+E35", "=F34+F35", "EBITDA + D&A", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Interest Income", 50,60,70,80,90, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Interest Expense", -800,-820,-850,-900,-950, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Other Income/Expense", 20,25,10,-5,15, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Profit Before Tax (PBT)", "=B36+B37+B38+B39", "=C36+C37+C38+C39", "=D36+D37+D38+D39", "=E36+E37+E38+E39", "=F36+F37+F38+F39", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Taxation",-460,-470,-460,-440,-460, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Profit After Tax", "=B40+B41", "=C40+C41", "=D40+D41", "=E40+E41", "=F40+F41", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Non-controlling Interests", -60,-65,-50,-55,-35, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Net Profit (for Equity Holders)", "=B42+B43", "=C42+C43", "=D42+D43", "=E42+E43", "=F42+F43", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN)
]
current_row = 2
for row_data_tuple in hpl_rows:
    desc, v1, v2, v3, v4, v5, notes_val, is_header_val, num_fmt_val = row_data_tuple[0], row_data_tuple[1], row_data_tuple[2], row_data_tuple[3], row_data_tuple[4], row_data_tuple[5], row_data_tuple[6], row_data_tuple[7], row_data_tuple[8]
    cell_A = ws_hpl.cell(row=current_row, column=1, value=desc)
    if is_header_val:
        ws_hpl.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(HISTORICAL_YEARS_DATA)+2)
        cell_A.font = FONT_HEADER; cell_A.fill = FILL_HEADER; cell_A.alignment = ALIGN_CENTER
    else:
        style_row_header(cell_A, level=2 if desc.startswith("Total") or "GROUP" in desc else 3)
        data_values = [v1,v2,v3,v4,v5]
        for i, val in enumerate(data_values):
            data_cell = ws_hpl.cell(row=current_row, column=i+2, value=val)
            style_data_cell(data_cell, is_input=not (isinstance(val, str) and val.startswith("=")), number_format=num_fmt_val)
        ws_hpl.cell(row=current_row, column=len(HISTORICAL_YEARS_DATA)+2, value=notes_val).border = BORDER_THIN_ALL
    current_row +=1

# --- Sheet: Hist_BS_Consol ---
ws_hbs = wb.create_sheet("Hist_BS_Consol")
set_column_widths(ws_hbs, {'A': 40, **{get_column_letter(i+2): 12 for i in range(len(HISTORICAL_YEARS_DATA))}, get_column_letter(len(HISTORICAL_YEARS_DATA)+2): 40})
setup_sheet_headers(ws_hbs, "Historical Balance Sheet (£m)", HISTORICAL_YEARS_DATA)
hbs_rows = [ # (Description, FY20, FY21, FY22, FY23, FY24, Notes, Is_Header, Num_Format)
    ("ASSETS", None, None, None, None, None, "", True, None),
    ("Non-Current Assets", None, None, None, None, None, "", True, None),
    ("Property, Plant & Equipment", 40000, 42000, 50000, 53000, 56000, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Intangible Assets", 8000, 8500, 9000, 9500, 10000, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Investments (JVs, Assoc.)", 1000, 1100, 1200, 1300, 1400, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Deferred Tax Assets", 500, 550, 600, 650, 700, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Other Non-Current Assets", 700, 750, 800, 850, 900, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Total Non-Current Assets", "=SUM(B4:B8)", "=SUM(C4:C8)", "=SUM(D4:D8)", "=SUM(E4:E8)", "=SUM(F4:F8)", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Current Assets", None, None, None, None, None, "", True, None),
    ("Inventories", 300, 320, 350, 380, 400, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Trade & Other Receivables", 2500, 2600, 2800, 3000, 3200, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Cash & Cash Equivalents", 1000, 1200, 1000, 1500, 1300, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Other Current Assets", 200, 210, 220, 230, 240, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Total Current Assets", "=SUM(B11:B14)", "=SUM(C11:C14)", "=SUM(D11:D14)", "=SUM(E11:E14)", "=SUM(F11:F14)", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("TOTAL ASSETS", "=B9+B15", "=C9+C15", "=D9+D15", "=E9+E15", "=F9+F15", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("LIABILITIES & EQUITY", None, None, None, None, None, "", True, None),
    ("Equity", None, None, None, None, None, "", True, None),
    ("Share Capital", 700, 700, 700, 700, 700, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Share Premium / Reserves", 10000, 10500, 11000, 11500, 12000, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Non-Controlling Interests", 300, 320, 350, 380, 400, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Total Equity", "=SUM(B19:B21)", "=SUM(C19:C21)", "=SUM(D19:D21)", "=SUM(E19:E21)", "=SUM(F19:F21)", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Non-Current Liabilities", None, None, None, None, None, "", True, None),
    ("Borrowings (Long-term)", 28000, 29500, 35000, 37000, 39000, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Deferred Tax Liabilities", 3000, 3200, 3500, 3700, 3900, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Provisions", 1000, 1050, 1100, 1150, 1200, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Other Non-Current Liab.", 800, 850, 900, 950, 1000, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Total Non-Current Liab.", "=SUM(B24:B27)", "=SUM(C24:C27)", "=SUM(D24:D27)", "=SUM(E24:E27)", "=SUM(F24:F27)", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Current Liabilities", None, None, None, None, None, "", True, None),
    ("Borrowings (Short-term)", 2000, 2500, 3000, 3500, 4000, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Trade & Other Payables", 3000, 3100, 3200, 3300, 3400, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Current Tax Liabilities", 500, 550, 600, 650, 700, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Other Current Liabilities", 4900, 5460, 7620, 8580, 8840, "Plug in example, ensure BS balances", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Total Current Liabilities", "=SUM(B30:B33)", "=SUM(C30:C33)", "=SUM(D30:D33)", "=SUM(E30:E33)", "=SUM(F30:F33)", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Total Liabilities", "=B28+B34", "=C28+C34", "=D28+D34", "=E28+E34", "=F28+F34", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("TOTAL LIABILITIES & EQUITY", "=B22+B35", "=C22+C35", "=D22+D35", "=E22+E35", "=F22+F35", "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Balance Check (Assets - L&E)", "=B16-B36", "=C16-C36", "=D16-D36", "=E16-E36", "=F16-F36", "Should be 0", False, FORMAT_NUMBER_0DP_NEG_PAREN)
]
current_row = 2
for row_data_tuple in hbs_rows:
    desc, v1, v2, v3, v4, v5, notes_val, is_header_val, num_fmt_val = row_data_tuple[0], row_data_tuple[1], row_data_tuple[2], row_data_tuple[3], row_data_tuple[4], row_data_tuple[5], row_data_tuple[6], row_data_tuple[7], row_data_tuple[8]
    cell_A = ws_hbs.cell(row=current_row, column=1, value=desc)
    if is_header_val:
        ws_hbs.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(HISTORICAL_YEARS_DATA)+2)
        cell_A.font = FONT_HEADER; cell_A.fill = FILL_HEADER; cell_A.alignment = ALIGN_CENTER
    else:
        style_row_header(cell_A, level=2 if desc.startswith("Total") or "TOTAL" in desc or "Balance Check" in desc else 3)
        data_values = [v1,v2,v3,v4,v5]
        for i, val in enumerate(data_values):
            data_cell = ws_hbs.cell(row=current_row, column=i+2, value=val)
            style_data_cell(data_cell, is_input=not (isinstance(val, str) and val.startswith("=")), number_format=num_fmt_val)
        ws_hbs.cell(row=current_row, column=len(HISTORICAL_YEARS_DATA)+2, value=notes_val).border = BORDER_THIN_ALL
    current_row +=1

# --- Sheet: Hist_CF_Consol ---
ws_hcf = wb.create_sheet("Hist_CF_Consol")
set_column_widths(ws_hcf, {'A': 45, **{get_column_letter(i+2): 12 for i in range(len(HISTORICAL_YEARS_DATA))}, get_column_letter(len(HISTORICAL_YEARS_DATA)+2): 40})
setup_sheet_headers(ws_hcf, "Historical Cash Flow (£m)", HISTORICAL_YEARS_DATA)
hcf_rows = [ # (Description, FY20, FY21, FY22, FY23, FY24, Notes, Is_Header, Num_Format)
    ("Cash Flow from Operating Activities (CFO)", None, None, None, None, None, "", True, None),
    ("Profit Before Tax", 2320,2335,2310,2195,2295, "From P&L", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Depreciation & Amortization", 1650,1730,1960,2490,2590, "From P&L", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Net CFO", 4170, 4295, 4650, 4945, 5315, "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Cash Flow from Investing Activities (CFI)", None, None, None, None, None, "", True, None),
    ("Purchase of PP&E (Capex)", -3500, -3800, -4500, -5000, -5500, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Net CFI", -3580, -3865, -9700, -4085, -5720, "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Cash Flow from Financing Activities (CFF)", None, None, None, None, None, "", True, None),
    ("Proceeds from Borrowings", 2000,2200,6000,1500,2000, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Repayment of Borrowings", -1000,-1200,-800,-1000,-1100, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Dividends Paid to Equity Holders", -1200, -1230, -1250, -1280, -1300, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Net CFF", -1030, -1085, 3130, -1695, -1305, "Calculated", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Net Change in Cash & Equivalents", -440, -655, -1920, -835, -1710, "CFO+CFI+CFF", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Cash at Beginning of Year", 1440, 1000, 345, -1575, -2410, "From prior year BS", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Cash at End of Year", 1000, 345, -1575, -2410, -4120, "Should match BS Cash", False, FORMAT_NUMBER_0DP_NEG_PAREN)
]
current_row = 2
for row_data_tuple in hcf_rows:
    desc, v1, v2, v3, v4, v5, notes_val, is_header_val, num_fmt_val = row_data_tuple[0], row_data_tuple[1], row_data_tuple[2], row_data_tuple[3], row_data_tuple[4], row_data_tuple[5], row_data_tuple[6], row_data_tuple[7], row_data_tuple[8]
    cell_A = ws_hcf.cell(row=current_row, column=1, value=desc)
    if is_header_val:
        ws_hcf.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(HISTORICAL_YEARS_DATA)+2)
        cell_A.font = FONT_HEADER; cell_A.fill = FILL_HEADER; cell_A.alignment = ALIGN_CENTER
    else:
        style_row_header(cell_A, level=2 if "Net C" in desc or "Cash at" in desc else 3)
        for i, val in enumerate(data_vals):
            data_cell = ws_hcf.cell(row=current_row, column=i+2, value=val)
            style_data_cell(data_cell, is_input=not (isinstance(val, str) and val.startswith("=")), number_format=num_fmt_val)
        ws_hcf.cell(row=current_row, column=len(HISTORICAL_YEARS_DATA)+2, value=notes_val).border = BORDER_THIN_ALL
    current_row +=1

# --- Sheet: Hist_RAV_RateBase ---
ws_hrav = wb.create_sheet("Hist_RAV_RateBase")
set_column_widths(ws_hrav, {'A': 45, **{get_column_letter(i+2): 12 for i in range(len(HISTORICAL_YEARS_DATA))}, get_column_letter(len(HISTORICAL_YEARS_DATA)+2): 40})
setup_sheet_headers(ws_hrav, "Historical RAV & Rate Base (£m or $m)", HISTORICAL_YEARS_DATA)
hrav_rows = [ # (Description, FY20, FY21, FY22, FY23, FY24, Notes, Is_Header, Num_Format)
    ("UK Electricity Transmission (NGET) - RAV", None, None, None, None, None, "£m", True, None),
    ("Opening RAV", 18000, 18800, 19700, 20800, 22000, "", False, FORMAT_NUMBER_0DP),
    ("Capex Additions (Allowed)", 1200, 1300, 1400, 1500, 1600, "", False, FORMAT_NUMBER_0DP),
    ("Regulatory Depreciation", -500, -520, -540, -560, -580, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Inflation Adjustment", 100, 120, 240, 260, 280, "", False, FORMAT_NUMBER_0DP),
    ("Closing RAV", "=SUM(B3:B6)", "=SUM(C3:C6)", "=SUM(D3:D6)", "=SUM(E3:E6)", "=SUM(F3:F6)", "Calculated", False, FORMAT_NUMBER_0DP),
    ("UK Electricity Distribution (NGED) - RAV", None, None, None, None, None, "£m", True, None),
    ("Opening RAV", 0,0,12000,13000,14000, "Acquired mid-FY22", False, FORMAT_NUMBER_0DP),
    ("Capex Additions (Allowed)", 0,0,500,1200,1300, "", False, FORMAT_NUMBER_0DP),
    ("Regulatory Depreciation", 0,0,-200,-500,-550, "", False, FORMAT_NUMBER_0DP_NEG_PAREN),
    ("Inflation Adjustment", 0,0,100,300,320, "", False, FORMAT_NUMBER_0DP),
    ("Closing RAV", "=SUM(B9:B12)", "=SUM(C9:C12)", "=SUM(D9:D12)", "=SUM(E9:E12)", "=SUM(F9:F12)", "Calculated", False, FORMAT_NUMBER_0DP),
    ("US Regulated - Rate Base (NY)", None, None, None, None, None, "$m (unless noted)", True, None),
    ("Closing Rate Base ($m)", 10500,11000,11500,12000,12500, "Illustrative", False, FORMAT_NUMBER_0DP),
    ("US Regulated - Rate Base (MA)", None, None, None, None, None, "$m (unless noted)", True, None),
    ("Closing Rate Base ($m)", 8300,8600,8900,9200,9500, "Illustrative", False, FORMAT_NUMBER_0DP),
]
current_row = 2
for row_data_tuple in hrav_rows:
    desc, v1, v2, v3, v4, v5, notes_val, is_header_val, num_fmt_val = row_data_tuple[0], row_data_tuple[1], row_data_tuple[2], row_data_tuple[3], row_data_tuple[4], row_data_tuple[5], row_data_tuple[6], row_data_tuple[7], row_data_tuple[8]
    cell_A = ws_hrav.cell(row=current_row, column=1, value=desc)
    if is_header_val:
        ws_hrav.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(HISTORICAL_YEARS_DATA)+2)
        cell_A.font = FONT_HEADER; cell_A.fill = FILL_HEADER; cell_A.alignment = ALIGN_CENTER
    else:
        style_row_header(cell_A, level=2 if "Closing" in desc else 3)
        for i, val in enumerate(data_vals):
            data_cell = ws_hrav.cell(row=current_row, column=i+2, value=val)
            style_data_cell(data_cell, is_input=not (isinstance(val, str) and val.startswith("=")), number_format=num_fmt_val)
        ws_hrav.cell(row=current_row, column=len(HISTORICAL_YEARS_DATA)+2, value=notes_val).border = BORDER_THIN_ALL
    current_row +=1

# --- Sheet: Assumptions_UK_Reg ---
ws_ukr = wb.create_sheet("Assumptions_UK_Reg")
set_column_widths(ws_ukr, {'A': 45, **{get_column_letter(i+2): 12 for i in range(len(FORECAST_YEARS_MODEL))}, get_column_letter(len(FORECAST_YEARS_MODEL)+2): 50})
setup_sheet_headers(ws_ukr, "UK Regulated Assumptions", FORECAST_YEARS_MODEL)
uk_reg_data = [
    ("NGET (RIIO-T2/T3)", None, None, "", True),
    ("RAV: Capex Additions (£m)", [1700, 1800, 1900, 2000, 2100] + [2200]*11, FORMAT_NUMBER_0DP, "Net of contribs. From investment plans."),
    ("RAV: Regulatory Depn Rate (% Opening RAV)", [0.025]*16, FORMAT_PERCENT_1DP, "Or abs £m. From Ofgem/company."),
    ("RAV: Inflation Link (CPIH Ref)", ["=Assumptions_Macro!C2"]*16, FORMAT_PERCENT_1DP, "='Assumptions_Macro'!C2 (drag right)"),
    ("Revenue: Allowed WACC (Nominal %)", [0.050, 0.050, 0.050, 0.048, 0.048] + [0.048]*11, FORMAT_PERCENT_1DP, "Illustrative. From Ofgem RIIO-T2/3."),
    ("Revenue: Outperformance/Underperformance (£m)", [50, 50, 25, 25, 0] + [0]*11, FORMAT_NUMBER_0DP_NEG_PAREN, "Net incentive earnings"),
    ("Opex: Base before efficiency (£m)", [1250, 1270, 1290, 1310, 1330] + [1350]*11, FORMAT_NUMBER_0DP_NEG_PAREN, "Grows with inflation & activity"),
    ("Opex: Efficiency Target (% reduction on base)", [0.01, 0.01, 0.005, 0.005, 0.005] + [0.005]*11, FORMAT_PERCENT_1DP, "Annual efficiency"),
    ("NGED (RIIO-ED2/ED3)", None, None, "", True),
    ("RAV: Capex Additions (£m)", [1400, 1500, 1600, 1700, 1800] + [1900]*11, FORMAT_NUMBER_0DP, "From investment plans."),
    ("RAV: Regulatory Depn Rate (% Opening RAV)", [0.030]*16, FORMAT_PERCENT_1DP, "From Ofgem/company."),
    ("RAV: Inflation Link (CPIH Ref)", ["=Assumptions_Macro!C2"]*16, FORMAT_PERCENT_1DP, "='Assumptions_Macro'!C2 (drag right)"),
    ("Revenue: Allowed WACC (Nominal %)", [0.048, 0.048, 0.048, 0.046, 0.046] + [0.046]*11, FORMAT_PERCENT_1DP, "Illustrative. From Ofgem RIIO-ED2/3."),
    ("Revenue: Outperformance/Underperformance (£m)", [40, 40, 20, 20, 0] + [0]*11, FORMAT_NUMBER_0DP_NEG_PAREN, "Net incentive earnings"),
    ("Opex: Base before efficiency (£m)", [1300, 1320, 1340, 1360, 1380] + [1400]*11, FORMAT_NUMBER_0DP_NEG_PAREN, "Grows with inflation & activity"),
    ("Opex: Efficiency Target (% reduction on base)", [0.01, 0.01, 0.005, 0.005, 0.005] + [0.005]*11, FORMAT_PERCENT_1DP, "Annual efficiency"),
]
current_row = 2
for item_data in uk_reg_data:
    item, values, num_format, notes_text = item_data[0], item_data[1], item_data[2], item_data[3]
    is_section_header = item_data[4] if len(item_data) > 4 else (values is None)
    cell_A = ws_ukr.cell(row=current_row, column=1, value=item)
    if is_section_header:
        ws_ukr.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(FORECAST_YEARS_MODEL)+2)
        cell_A.font = FONT_HEADER; cell_A.fill = FILL_HEADER; cell_A.alignment = ALIGN_CENTER
    else:
        style_row_header(cell_A, level=2)
        ws_ukr.cell(row=current_row, column=len(FORECAST_YEARS_MODEL)+2, value=notes_text).border = BORDER_THIN_ALL
        for i, val in enumerate(values):
            data_cell = ws_ukr.cell(row=current_row, column=i+2, value=val)
            if isinstance(val, str) and val.startswith("="): # Formula
                style_data_cell(data_cell, is_input=False, number_format=num_format)
            else: # Input
                style_data_cell(data_cell, is_input=True, number_format=num_format)
    current_row += 1

# --- Sheet: Assumptions_US_Reg ---
ws_usr = wb.create_sheet("Assumptions_US_Reg")
set_column_widths(ws_usr, {'A': 45, **{get_column_letter(i+2): 12 for i in range(len(FORECAST_YEARS_MODEL))}, get_column_letter(len(FORECAST_YEARS_MODEL)+2): 50})
setup_sheet_headers(ws_usr, "US Regulated Assumptions", FORECAST_YEARS_MODEL)
us_reg_data = [
    ("New York (NY)", None, None, "", True),
    ("Rate Base: Capex Additions ($m)", [1200, 1250, 1300, 1350, 1400] + [1450]*11, FORMAT_NUMBER_0DP, "Original currency. Company plans."),
    ("Rate Base: Book Depn Rate (% Opening RB)", [0.035]*16, FORMAT_PERCENT_1DP, "Or $m. From PUC filings."),
    ("Revenue: Allowed ROE (%)", [0.090, 0.090, 0.091, 0.091, 0.092] + [0.092]*11, FORMAT_PERCENT_1DP, "From latest rate case."),
    ("Revenue: Equity Ratio in Cap Structure (%)", [0.50]*16, FORMAT_PERCENT_0DP, "For ratemaking purposes."),
    ("Revenue: Overall Growth Rate (%)", [0.040,0.042,0.045,0.040,0.038] + [0.035]*11, FORMAT_PERCENT_1DP, "Incl. rate changes & underlying growth."),
    ("Opex: Growth (before US CPI inflation) (%)", [0.01, 0.01, 0.008, 0.005, 0.005] + [0.005]*11, FORMAT_PERCENT_1DP, "Underlying opex growth."),
    ("Massachusetts (MA)", None, None, "", True),
    ("Rate Base: Capex Additions ($m)", [900, 950, 1000, 1050, 1100] + [1150]*11, FORMAT_NUMBER_0DP, "Original currency."),
    ("Rate Base: Book Depn Rate (% Opening RB)", [0.033]*16, FORMAT_PERCENT_1DP, ""),
    ("Revenue: Allowed ROE (%)", [0.092, 0.092, 0.093, 0.093, 0.094] + [0.094]*11, FORMAT_PERCENT_1DP, ""),
    ("Revenue: Equity Ratio in Cap Structure (%)", [0.52]*16, FORMAT_PERCENT_0DP, ""),
    ("Revenue: Overall Growth Rate (%)", [0.038,0.040,0.042,0.038,0.036] + [0.033]*11, FORMAT_PERCENT_1DP, ""),
    ("Opex: Growth (before US CPI inflation) (%)", [0.008, 0.008, 0.006, 0.004, 0.004] + [0.004]*11, FORMAT_PERCENT_1DP, ""),
]
current_row = 2
for item_data in us_reg_data:
    item, values, num_format, notes_text = item_data[0], item_data[1], item_data[2], item_data[3]
    is_section_header = item_data[4] if len(item_data) > 4 else (values is None)
    cell_A = ws_usr.cell(row=current_row, column=1, value=item)
    if is_section_header:
        ws_usr.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(FORECAST_YEARS_MODEL)+2)
        cell_A.font = FONT_HEADER; cell_A.fill = FILL_HEADER; cell_A.alignment = ALIGN_CENTER
    else:
        style_row_header(cell_A, level=2)
        ws_usr.cell(row=current_row, column=len(FORECAST_YEARS_MODEL)+2, value=notes_text).border = BORDER_THIN_ALL
        for i, val in enumerate(values):
            data_cell = ws_usr.cell(row=current_row, column=i+2, value=val)
            style_data_cell(data_cell, is_input=True, number_format=num_format)
    current_row += 1

# --- Sheet: Assumptions_NGV ---
ws_ngva = wb.create_sheet("Assumptions_NGV")
set_column_widths(ws_ngva, {'A': 45, **{get_column_letter(i+2): 12 for i in range(len(FORECAST_YEARS_MODEL))}, get_column_letter(len(FORECAST_YEARS_MODEL)+2): 50})
setup_sheet_headers(ws_ngva, "NGV Assumptions", FORECAST_YEARS_MODEL)
ngv_data = [
    ("INTERCONNECTORS", None, None, "", True),
    ("IFA1/2 Revenue (£m)", [158,191,179,161,146] + [140]*11, FORMAT_NUMBER_0DP_NEG_PAREN, "Capacity * Avail * Spread * Hours"),
    ("IFA1/2 Opex (£m)", [-30,-31,-32,-33,-34] + [-35]*11, FORMAT_NUMBER_0DP_NEG_PAREN, "Grows with inflation"),
    ("BritNed Revenue (£m)", [96,106,90,82,82] + [80]*11, FORMAT_NUMBER_0DP_NEG_PAREN, ""),
    ("BritNed Opex (£m)", [-20,-21,-22,-23,-24] + [-25]*11, FORMAT_NUMBER_0DP_NEG_PAREN, ""),
    ("Total Interconnectors Capex (£m)", [-30,-30,-70,-70,-30] + [-25]*11, FORMAT_NUMBER_0DP_NEG_PAREN, "Sum of project capex"),
    ("GRAIN LNG", None, None, "", True),
    ("Grain LNG Revenue (£m)", [150,153,146,148,140] + [135]*11, FORMAT_NUMBER_0DP_NEG_PAREN, "Contracted capacity * tariff"),
    ("Grain LNG Opex (£m)", [-50,-51,-52,-53,-54] + [-55]*11, FORMAT_NUMBER_0DP_NEG_PAREN, "Grows with inflation"),
    ("Grain LNG Capex (£m)", [-20,-30,-40,-20,-15] + [-10]*11, FORMAT_NUMBER_0DP_NEG_PAREN, "Expansion/Maintenance")
]
current_row = 2
for item_data in ngv_data:
    item, values, num_format, notes_text = item_data[0], item_data[1], item_data[2], item_data[3]
    is_section_header = item_data[4] if len(item_data) > 4 else (values is None)
    cell_A = ws_ngva.cell(row=current_row, column=1, value=item)
    if is_section_header:
        ws_ngva.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(FORECAST_YEARS_MODEL)+2)
        cell_A.font = FONT_HEADER; cell_A.fill = FILL_HEADER; cell_A.alignment = ALIGN_CENTER
    else:
        style_row_header(cell_A, level=2)
        ws_ngva.cell(row=current_row, column=len(FORECAST_YEARS_MODEL)+2, value=notes_text).border = BORDER_THIN_ALL
        for i, val in enumerate(values):
            data_cell = ws_ngva.cell(row=current_row, column=i+2, value=val)
            style_data_cell(data_cell, is_input=True, number_format=num_format)
    current_row += 1

# --- Sheet: Forecast_PL_Segment ---
ws_fpl = wb.create_sheet("Forecast_PL_Segment")
set_column_widths(ws_fpl, {'A': 40, **{get_column_letter(i+2): 12 for i in range(len(DISPLAY_YEARS))}, get_column_letter(len(DISPLAY_YEARS)+2): 70})
setup_sheet_headers(ws_fpl, "Forecast P&L by Segment (£m)", DISPLAY_YEARS)
# ... (Formula logic for each row and year needs to be meticulously implemented here)
# This involves linking to Hist_PL_Segment for FY23, FY24 and then using Assumptions for FY25 onwards.

# --- Sheet: RAV_RateBase_Forecast ---
ws_frav = wb.create_sheet("RAV_RateBase_Forecast")
set_column_widths(ws_frav, {'A': 45, **{get_column_letter(i+2): 12 for i in range(len(FORECAST_YEARS_MODEL))}, get_column_letter(len(FORECAST_YEARS_MODEL)+2): 70})
setup_sheet_headers(ws_frav, "Forecast RAV & Rate Base", FORECAST_YEARS_MODEL, notes_col=True)
# ... (Formulas for RAV/Rate Base roll forward)

# --- Sheet: Debt_Schedule_Forecast ---
ws_fdebt = wb.create_sheet("Debt_Schedule_Forecast")
set_column_widths(ws_fdebt, {'A': 40, **{get_column_letter(i+2): 12 for i in range(len(FORECAST_YEARS_MODEL))}, get_column_letter(len(FORECAST_YEARS_MODEL)+2): 70})
setup_sheet_headers(ws_fdebt, "Forecast Debt Schedule (£m)", FORECAST_YEARS_MODEL)
# ... (Formulas for debt roll forward, interest calculations, debt plug link from CF)

# --- Sheet: Forecast_CF_Consol ---
ws_fcf = wb.create_sheet("Forecast_CF_Consol")
set_column_widths(ws_fcf, {'A': 45, **{get_column_letter(i+2): 12 for i in range(len(DISPLAY_YEARS))}, get_column_letter(len(DISPLAY_YEARS)+2): 70})
setup_sheet_headers(ws_fcf, "Forecast Cash Flow (£m)", DISPLAY_YEARS)
# ... (Formulas linking to P&L, BS, Assumptions, Debt Schedule; calculates debt plug)

# --- Sheet: Forecast_BS_Consol ---
ws_fbs = wb.create_sheet("Forecast_BS_Consol")
set_column_widths(ws_fbs, {'A': 40, **{get_column_letter(i+2): 12 for i in range(len(DISPLAY_YEARS))}, get_column_letter(len(DISPLAY_YEARS)+2): 70})
setup_sheet_headers(ws_fbs, "Forecast Balance Sheet (£m)", DISPLAY_YEARS)
# ... (Formulas linking to Hist_BS, P&L, CF, Debt, RAV. Includes Balance Check row)

# --- Sheet: Credit_Metrics ---
ws_cred = wb.create_sheet("Credit_Metrics")
set_column_widths(ws_cred, {'A': 35, **{get_column_letter(i+2): 12 for i in range(len(DISPLAY_YEARS))}, get_column_letter(len(DISPLAY_YEARS)+2): 60})
setup_sheet_headers(ws_cred, "Credit Metrics", DISPLAY_YEARS)
# ... (Formulas linking to P&L, BS, CF to calculate key ratios)

# --- Sheet: Cover_Summary ---
ws_summ = wb.create_sheet("Cover_Summary")
summary_display_cols = HISTORICAL_YEARS_DATA[-1:] + [FORECAST_YEARS_MODEL[0], FORECAST_YEARS_MODEL[1], FORECAST_YEARS_MODEL[2], FORECAST_YEARS_MODEL[5], FORECAST_YEARS_MODEL[10], FORECAST_YEARS_MODEL[-1]]
set_column_widths(ws_summ, {'A': 40, **{get_column_letter(i+2): 14 for i in range(len(summary_display_cols))}})
setup_sheet_headers(ws_summ, "Model Summary", summary_display_cols, notes_col=False, main_header_fill=FILL_GREY, year_header_fill=FILL_GREY)
# ... (Direct links to key outputs from other forecast sheets)

# Move Cover_Summary to be the first sheet
if "Cover_Summary" in wb.sheetnames: # Check if sheets exist
    wb.move_sheet(ws_summ, offset=-len(wb.sheetnames)+1)

# --- Final Save ---
output_filename = "NationalGrid_FinancialModel_Generated.xlsx"
try:
    wb.save(output_filename)
    # print(f"Successfully created '{output_filename}'") # Cannot use print in this environment
except Exception as e:
    # print(f"Error saving workbook: {e}") # Cannot use print
    pass # Placeholder for error logging or handling
