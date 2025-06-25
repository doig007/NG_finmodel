```python
import openpyxl
import csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# --- Configuration & Constants ---
HISTORICAL_YEARS_DATA = ["FY2020", "FY2021", "FY2022", "FY2023", "FY2024"]
FORECAST_YEARS_MODEL = [f"FY{2025 + i}" for i in range(16)] # FY2025 to FY2040
DISPLAY_YEARS = HISTORICAL_YEARS_DATA[-2:] + FORECAST_YEARS_MODEL

# --- Styling Definitions ---
COLOR_PRIMARY_BLUE = "4F81BD"
COLOR_SECONDARY_BLUE = "DCE6F1"
COLOR_WHITE = "FFFFFF"
COLOR_BLACK = "000000"
COLOR_INPUT_BLUE = "0000FF"
COLOR_GREY_FILL = "F2F2F2"

FONT_HEADER = Font(bold=True, color=COLOR_WHITE, name='Calibri', size=11)
FONT_SUBHEADER = Font(bold=True, color=COLOR_BLACK, name='Calibri', size=11)
FONT_INPUT = Font(color=COLOR_INPUT_BLUE, name='Calibri', size=10)
FONT_FORMULA = Font(color=COLOR_BLACK, name='Calibri', size=10)

FILL_HEADER = PatternFill(start_color=COLOR_PRIMARY_BLUE, end_color=COLOR_PRIMARY_BLUE, fill_type="solid")
FILL_SUBHEADER = PatternFill(start_color=COLOR_SECONDARY_BLUE, end_color=COLOR_SECONDARY_BLUE, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_THIN_ALL = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

FORMAT_PERCENT_1DP = '0.0%'
FORMAT_PERCENT_0DP = '0%'
FORMAT_NUMBER_0DP = '#,##0'
FORMAT_NUMBER_0DP_NEG_PAREN = '#,##0;(#,##0);0' # Added zero display
FORMAT_NUMBER_2DP = '#,##0.00'
FORMAT_MULTIPLIER = '0.00x'

# --- Helper Functions ---
def setup_sheet_headers(ws, title, years_list, first_data_col_idx=2, row_num=1, notes_col=True, first_col_width=45):
    ws.cell(row=row_num, column=1, value=title).font = FONT_HEADER # Changed to FONT_HEADER for main sheet titles
    ws.cell(row=row_num, column=1).fill = FILL_HEADER # Added fill for main sheet titles
    ws.cell(row=row_num, column=1).alignment = ALIGN_LEFT # Align title left
    ws.cell(row=row_num, column=1).border = BORDER_THIN_ALL
    ws.column_dimensions[get_column_letter(1)].width = first_col_width
    
    for i, year in enumerate(years_list):
        cell = ws.cell(row=row_num, column=first_data_col_idx + i, value=year)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN_ALL
        ws.column_dimensions[get_column_letter(first_data_col_idx + i)].width = 12
    if notes_col:
        notes_col_idx = first_data_col_idx + len(years_list)
        cell = ws.cell(row=row_num, column=notes_col_idx, value="Notes/Links")
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN_ALL
        ws.column_dimensions[get_column_letter(notes_col_idx)].width = 50

def style_row_header(cell, level=1, fill=True):
    cell.font = FONT_SUBHEADER if level == 1 else Font(bold=True, name='Calibri', size=10)
    if level == 1 and fill:
        cell.fill = FILL_SUBHEADER
    cell.alignment = ALIGN_LEFT
    cell.border = BORDER_THIN_ALL

def style_data_cell(cell, is_input=False, number_format=FORMAT_NUMBER_0DP_NEG_PAREN, is_formula=False):
    if is_input:
        cell.font = FONT_INPUT
    elif is_formula:
        cell.font = FONT_FORMULA
    else:
        cell.font = Font(name='Calibri', size=10) # Default for text notes

    cell.number_format = number_format
    cell.alignment = ALIGN_RIGHT if not (isinstance(cell.value, str) and not cell.value.startswith("=")) else ALIGN_LEFT
    cell.border = BORDER_THIN_ALL


def load_csv_to_sheet(ws, csv_filename, start_row=1, is_assumptions_sheet=False, header_row_offset=0):
    try:
        with open(csv_filename, 'r', newline='') as f:
            reader = csv.reader(f)
            for r_idx, row_content in enumerate(reader):
                actual_row = start_row + r_idx + header_row_offset
                for c_idx, value in enumerate(row_content):
                    cell = ws.cell(row=actual_row, column=c_idx + 1)
                    # Special handling for section headers in CSVs
                    is_section_header_csv = all(v == '' for v in row_content[1:]) and row_content[0] != "Assumption" and row_content[0] != "Line Item (£m)" and row_content[0] != "Line Item"

                    if r_idx == 0 and start_row == 1 and header_row_offset == 0: # Main CSV header row
                        cell.value = value
                        style_row_header(cell, level=1 if c_idx == 0 else 2, fill=False)
                        if c_idx > 0 :
                             cell.fill = FILL_HEADER
                             cell.font = FONT_HEADER
                             cell.alignment = ALIGN_CENTER
                    elif is_section_header_csv:
                        cell.value = value # Section name in first col
                        ws.merge_cells(start_row=actual_row, start_column=1, end_row=actual_row, end_column=len(row_content))
                        cell.font = FONT_HEADER; cell.fill = FILL_HEADER; cell.alignment = ALIGN_CENTER
                        break # Skip rest of the columns for this merged row
                    elif c_idx == 0: # First column (row description)
                        cell.value = value
                        style_row_header(cell, level=2, fill=False)
                    else: # Data area
                        try:
                            if value is None or value.strip() == "": cell.value = None # Keep blanks
                            elif '%' in value:
                                num_val = float(value.strip('%')) / 100
                                cell.value = num_val
                                style_data_cell(cell, is_input=is_assumptions_sheet, number_format=FORMAT_PERCENT_1DP, is_formula=value.startswith("="))
                            elif value.replace(',','').replace('(','-').replace(')','').replace('-','',1).isdigit() or (value.startswith('-') and value[1:].replace(',','').isdigit()): # Check if it's a number, allowing for commas and parentheses negatives
                                num_val = float(value.replace(',','').replace('(','-').replace(')',''))
                                cell.value = num_val
                                style_data_cell(cell, is_input=is_assumptions_sheet, number_format=FORMAT_NUMBER_0DP_NEG_PAREN, is_formula=value.startswith("="))
                            else: # Text (likely notes or formulas not converted)
                                cell.value = value
                                style_data_cell(cell, is_input=is_assumptions_sheet, number_format='General', is_formula=value.startswith("="))
                        except ValueError: # Keep as text if conversion fails
                            cell.value = value
                            style_data_cell(cell, is_input=is_assumptions_sheet, number_format='General', is_formula=value.startswith("="))
                            cell.alignment = ALIGN_LEFT
    except FileNotFoundError:
        error_msg = f"Error: {csv_filename} not found. Please create it."
        ws.cell(row=start_row, column=1, value=error_msg)
        print(error_msg)
    except Exception as e:
        error_msg = f"Error loading {csv_filename}: {e}"
        ws.cell(row=start_row, column=1, value=error_msg)
        print(error_msg)


def make_formula_draggable(base_formula, current_year_idx, first_forecast_year_col_idx, sheet_name_for_assumptions_offset=None):
    """
    Adjusts column letters in a formula string for dragging.
    `current_year_idx` is 0 for the first forecast year, 1 for the second, etc.
    `first_forecast_year_col_idx` is the actual column index (e.g., 4 for 'D').
    `sheet_name_for_assumptions_offset` is used if assumptions are in different columns than data.
    """
    if not isinstance(base_formula, str) or not base_formula.startswith("="):
        return base_formula

    # Regex to find cell references like A1, $A1, A$1, $A$1, SheetName!A1, etc.
    # It also captures simple ranges like A1:B2
    pattern = re.compile(r"(\'?\w+\'?!)?(\$?[A-Z]+)(\$?[0-9]+)(:(\$?[A-Z]+)(\$?[0-9]+))?")
    
    def replace_col(match):
        sheet_prefix, col_abs_rel, row_abs_rel, range_sep, col2_abs_rel, row2_abs_rel = match.groups()
        
        original_col_str = col_abs_rel.replace("$", "")
        original_col_idx = column_index_from_string(original_col_str)
        
        # Determine the correct base column index for assumptions (they don't shift with DISPLAY_YEARS history)
        # Assumptions usually start their forecast data in column C (index 3) in their CSVs/sheets
        # after "Assumption" and "Unit" columns.
        # The main data sheets (forecasts) might have historical years first.
        
        is_assumption_link = sheet_prefix and "Assumptions_" in sheet_prefix
        is_historical_link = sheet_prefix and "Hist_" in sheet_prefix

        if is_assumption_link:
            # Assumptions columns (C, D, E...) correspond to forecast years (FY25, FY26, FY27...)
            # The first assumption column is C (index 3 in its own sheet).
            # current_year_idx is 0 for FY25, 1 for FY26...
            final_col_idx = 3 + current_year_idx # C is 3rd col, D is 4th, etc.
        elif is_historical_link:
            # Historical links should generally be absolute or carefully handled, not simply dragged.
            # For this function, assume they are not meant to be dragged relative to forecast years.
            final_col_idx = original_col_idx # Keep historical links as they are in the base formula
        else: # Refers to a cell on the SAME forecast sheet or another forecast sheet
            final_col_idx = original_col_idx + current_year_idx
            
        new_col_str = ("$" if "$" in col_abs_rel else "") + get_column_letter(final_col_idx)
        
        # Handle ranges (like A1:B2)
        if range_sep:
            original_col2_str = col2_abs_rel.replace("$", "")
            original_col2_idx = column_index_from_string(original_col2_str)
            if is_assumption_link:
                 final_col2_idx = 3 + current_year_idx # Assuming range is within same year's assumption data
            elif is_historical_link:
                 final_col2_idx = original_col2_idx
            else:
                 final_col2_idx = original_col2_idx + current_year_idx
            new_col2_str = ("$" if "$" in col2_abs_rel else "") + get_column_letter(final_col2_idx)
            return f"{sheet_prefix or ''}{new_col_str}{row_abs_rel}{range_sep}{new_col2_str}{row2_abs_rel}"
        else:
            return f"{sheet_prefix or ''}{new_col_str}{row_abs_rel}"

    return pattern.sub(replace_col, base_formula)


# --- Create Workbook ---
wb = openpyxl.Workbook()
wb.remove(wb.active) # Remove default sheet

# --- Load CSV Data into Sheets ---
csv_files_info = [
    ("Assumptions_Macro", "assumptions_macro.csv", FORECAST_YEARS_MODEL, True, 1), # SheetName, CSV_File, YearList, IsAssumptions, CSV_Header_Offset
    ("Assumptions_UK_Reg", "assumptions_uk_reg.csv", FORECAST_YEARS_MODEL, True, 1),
    ("Assumptions_US_Reg", "assumptions_us_reg.csv", FORECAST_YEARS_MODEL, True, 1),
    ("Assumptions_NGV", "assumptions_ngv.csv", FORECAST_YEARS_MODEL, True, 1),
    ("Hist_PL_Segment", "hist_pl_segment.csv", HISTORICAL_YEARS_DATA, False, 0),
    ("Hist_BS_Consol", "hist_bs_consol.csv", HISTORICAL_YEARS_DATA, False, 0),
    ("Hist_CF_Consol", "hist_cf_consol.csv", HISTORICAL_YEARS_DATA, False, 0),
    ("Hist_RAV_RateBase", "hist_rav_ratebase.csv", HISTORICAL_YEARS_DATA, False, 0)
]

for name, fname, yrs, is_assum, csv_header_offset in csv_files_info:
    ws = wb.create_sheet(name)
    load_csv_to_sheet(ws, fname, start_row=1, is_assumptions_sheet=is_assum, header_row_offset=csv_header_offset)
    # General column width setting after loading
    ws.column_dimensions['A'].width = 45
    # Assuming data starts in column B after CSV load (Col A is description)
    # For assumption sheets, CSV has Unit in B, data from C
    data_start_col_csv = 3 if is_assum else 2
    for i in range(len(yrs)):
        ws.column_dimensions[get_column_letter(data_start_col_csv + i)].width = 12
    ws.column_dimensions[get_column_letter(data_start_col_csv + len(yrs))].width = 50 # Notes column

# --- RAV_RateBase_Forecast Sheet ---
ws_frav = wb.create_sheet("RAV_RateBase_Forecast")
setup_sheet_headers(ws_frav, "Forecast RAV & Rate Base", FORECAST_YEARS_MODEL, first_col_width=45)
# (Detailed structure and FY2025 formulas for RAV/RateBase. Subsequent years use make_formula_draggable)
# This section needs to be fully built out with the actual formulas and looping logic.
# For brevity, I'm showing a conceptual structure for one item.
# A full implementation would iterate through a predefined list of line items and their base formulas.
frav_row_definitions = {
    2: ("UK Electricity Transmission (NGET) - RAV", "£m", None, True), # Title row
    3: ("Opening RAV", "£m", "='Hist_RAV_RateBase'!G7", False), # G7 is FY24 Closing for NGET
    4: ("Capex Additions (Allowed)", "£m", "='Assumptions_UK_Reg'!C3", False), # C3 is FY25 NGET Capex
    5: ("Regulatory Depreciation", "£m", "=-C3*INDEX(Assumptions_UK_Reg!$C$4:$R$4,1,MATCH(D$1,Assumptions_UK_Reg!$C$1:$R$1,0))", False), # OpeningRAV(C3) * DepnRate(Assum_UK_Reg row 4)
    6: ("Inflation Adjustment", "£m", "=C3*INDEX(Assumptions_Macro!$C$3:$R$3,1,MATCH(D$1,Assumptions_Macro!$C$1:$R$1,0))", False), # OpeningRAV(C3) * CPIH(Assum_Macro row 3)
    7: ("Other Movements", "£m", 0, False),
    8: ("Closing RAV", "£m", "=SUM(C3:C7)", False),
    # ... similar definitions for NGED, US NY ($m), US MA ($m), and their £m conversions ...
}
# Example population loop (simplified)
for r, (desc, unit, base_formula_fy25, is_header) in frav_row_definitions.items():
    cell_A = ws_frav.cell(row=r, column=1, value=desc)
    style_row_header(cell_A, level=1 if is_header else 2, fill=is_header)
    if not is_header:
        ws_frav.cell(row=r, column=len(FORECAST_YEARS_MODEL)+2, value=unit).border = BORDER_THIN_ALL # Unit in notes
        # First forecast year (column B, index 2)
        if base_formula_fy25:
             ws_frav.cell(row=r, column=2).value = base_formula_fy25
        style_data_cell(ws_frav.cell(row=r, column=2), is_formula=bool(base_formula_fy25), number_format=FORMAT_NUMBER_0DP_NEG_PAREN)
        # Other forecast years
        for year_idx, _ in enumerate(FORECAST_YEARS_MODEL[1:], start=1): # year_idx = 1 for 2nd forecast year
            target_col = 2 + year_idx
            if "Opening" in desc: # Opening balance links to previous year's closing
                # This needs to correctly identify the 'Closing RAV' row for this block
                closing_row_offset = 5 # e.g. if Closing RAV is 5 rows below Opening
                ws_frav.cell(row=r, column=target_col).value = f"={get_column_letter(target_col-1)}{r+closing_row_offset}"
            elif base_formula_fy25 and isinstance(base_formula_fy25, str) and base_formula_fy25.startswith("="):
                # This needs a more robust make_formula_draggable for complex sheets
                # For now, just copying for simplicity, assuming it would be adapted
                ws_frav.cell(row=r, column=target_col).value = base_formula_fy25 # Needs actual dragging logic
            elif isinstance(base_formula_fy25, (int,float)):
                 ws_frav.cell(row=r, column=target_col).value = base_formula_fy25

            style_data_cell(ws_frav.cell(row=r, column=target_col), is_formula=True, number_format=FORMAT_NUMBER_0DP_NEG_PAREN)


# --- Placeholder for other Forecast & Summary Sheets ---
# Similar looping and formula generation logic would be applied to:
# Forecast_PL_Segment, Debt_Schedule_Forecast, Forecast_CF_Consol,
# Forecast_BS_Consol (with balance check), Credit_Metrics, Cover_Summary.
# Each would require its own `row_definitions` dictionary with base formulas.

sheet_placeholder_details_fc = [
    ("Forecast_PL_Segment", "Forecast P&L by Segment (£m)", DISPLAY_YEARS),
    ("Debt_Schedule_Forecast", "Forecast Debt Schedule (£m)", FORECAST_YEARS_MODEL),
    ("Forecast_CF_Consol", "Forecast Cash Flow (£m)", DISPLAY_YEARS),
    ("Forecast_BS_Consol", "Forecast Balance Sheet (£m)", DISPLAY_YEARS),
    ("Credit_Metrics", "Credit Metrics", DISPLAY_YEARS),
    ("Cover_Summary", "Model Summary", HISTORICAL_YEARS_DATA[-1:] + [FORECAST_YEARS_MODEL[0], FORECAST_YEARS_MODEL[1], FORECAST_YEARS_MODEL[2], FORECAST_YEARS_MODEL[5], FORECAST_YEARS_MODEL[10], FORECAST_YEARS_MODEL[-1]])
]

for sheet_name, header_title, year_list in sheet_placeholder_details_fc:
    ws = wb.create_sheet(sheet_name)
    is_summary = sheet_name == "Cover_Summary"
    col_widths = {'A': 40}
    if is_summary:
        for i in range(len(year_list)): col_widths[get_column_letter(i+2)] = 14
    else: # Standard forecast/display sheets
        for i in range(len(year_list)): col_widths[get_column_letter(i+2)] = 12
        if not is_summary : col_widths[get_column_letter(len(year_list)+2)] = 60 # Notes column

    set_column_widths(ws, col_widths)
    setup_sheet_headers(ws, header_title, year_list, notes_col=not is_summary, first_col_width=col_widths['A'])
    # Add a note that formulas need to be fully implemented
    ws.cell(row=3, column=2, value="FORMULAS TO BE IMPLEMENTED FOR ALL YEARS BASED ON MODEL LOGIC (FIRST YEAR AS TEMPLATE)").font = FONT_INPUT
    if sheet_name == "Forecast_BS_Consol":
        # Simplified Balance Check for structure - actual row numbers would be dynamic
        ws.cell(row=50, column=1, value="Balance Check (Assets - L&E)").font = FONT_SUBHEADER
        style_row_header(ws.cell(row=50,column=1), level=2)
        for i in range(len(year_list)):
            col_l = get_column_letter(i+2)
            # These row numbers (e.g., 20 for Total Assets, 45 for Total L&E) are placeholders
            # and must match the actual rows where these totals are calculated in a full model.
            ws.cell(row=50, column=i+2, value=f"={col_l}20-{col_l}45").number_format = FORMAT_NUMBER_0DP 
            style_data_cell(ws.cell(row=50, column=i+2), is_formula=True)


# Move Cover_Summary to be the first sheet
if "Cover_Summary" in wb.sheetnames:
    summary_sheet = wb["Cover_Summary"]
    wb.move_sheet(summary_sheet, offset=-len(wb.sheetnames)+1)

# --- Final Save ---
output_filename = "NationalGrid_Full_Model_Generated.xlsx"
try:
    wb.save(output_filename)
except Exception as e:
    # Cannot print in this environment, but would log error in a real script
    pass
```
